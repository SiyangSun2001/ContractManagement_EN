import Contract
import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

dateToday = datetime.today().strftime('%Y-%m-%d')
year = dateToday[0:4]
month = dateToday[5:7]
#get date for today for later use


def checkNewSheets(path):
    path = path + '\\' + 'SalesRecord'
    os.chdir(path)
    entries = os.listdir()
    setOfEntries = set(entries)
    #change directory to the sale record folder

    fileName = year + 'SalesRecord' + '.xlsx'
    #make sure there all sheets are already created, if not create new ones
    if fileName in setOfEntries:


        wb = openpyxl.load_workbook(fileName)
        sheets = wb.get_sheet_names()
        if (month + '月') not in sheets:
            wb.create_sheet(title=month + '月')
            pageSetup(wb, (month + '月'))
            wb.save(year + 'SalesRecord' + '.xlsx')
    else:
        wb = openpyxl.Workbook()
        sheet  = wb.active
        sheet.title = month + '月'
        pageSetup(wb, (month + '月'))
        wb.save(year + 'SalesRecord' + '.xlsx')



def enterData(path, listContract):
    path = path + '\\' + 'SalesRecord'
    os.chdir(path)
    wb = openpyxl.load_workbook(year + 'SalesRecord' + '.xlsx')
    sheet = wb.get_sheet_by_name(month + '月')

    #the following for loop finds the closest empty row in the sheet
    for i in sheet['A']:
        if i.value == None:
            max = int(i.row)
            break
        else:
            max = sheet.max_row + 1
    # keep count of each individual product within each entry
    A  = 0
    B = 0
    C = 0
    BX = 0
    BM = 0
    BS = 0
    XZ = 0

    for contract in listContract:
        for i in range(len(contract.price)):
            sheet.cell(row=max, column=1).value = dateToday
            sheet.cell(row=max, column=1).border = thin_border
            sheet.cell(row=max, column=2).value = contract.contractNum
            sheet.cell(row=max, column=2).border = thin_border
            sheet.cell(row=max, column=3).value = contract.companyName
            sheet.cell(row=max, column=3).border = thin_border
            sheet.cell(row=max, column=4).value = contract.modelNumber[i]
            sheet.cell(row=max, column=4).border = thin_border
            sheet.cell(row=max, column=4).alignment = Alignment(horizontal='left')
            sheet.cell(row=max, column=5).value = contract.modelCount[i]
            sheet.cell(row=max, column=5).border = thin_border
            sheet.cell(row=max, column=5).alignment = Alignment(horizontal='center')
            sheet.cell(row=max, column=6).value = contract.getFormattedPrice()[i]
            sheet.cell(row=max, column=6).border = thin_border
            sheet.cell(row=1, column=6).value += contract.price[i] * contract.modelCount[i]
            if 'A' in contract.modelNumber[i]:
                A += contract.modelCount[i]
            elif 'B' in contract.modelNumber[i] and 'BX' not in contract.modelNumber[i] and 'BX' not in contract.modelNumber[i] and 'BS' not in contract.modelNumber[i] :
                B += contract.modelCount[i]
            elif 'C' in contract.modelNumber[i]:
                C += contract.modelCount[i]
            elif 'BX' in contract.modelNumber[i] and 'BM' not in contract.modelNumber[i] and 'BS' not in contract.modelNumber[i]:
                BX += contract.modelCount[i]
            elif 'BM' in contract.modelNumber[i] and 'BS' not in contract.modelNumber[i]:
                BM += contract.modelCount[i]
            elif 'BS' in contract.modelNumber[i]:
                BS += contract.modelCount[i]
            elif 'XZ' in contract.modelNumber[i]:
                XZ += contract.modelCount[i]
            max += 1
    sheet['H3'] = sheet['H3'].value + A
    sheet['H4'] = sheet['H4'].value + B
    sheet['H5'] = sheet['H5'].value + BX
    sheet['H6'] = sheet['H6'].value + BM
    sheet['H7'] = sheet['H7'].value + BS
    sheet['H8'] = sheet['H8'].value + XZ
    sheet['H9'] = sheet['H9'].value + C
    sheet['H2'] = sheet['H3'].value + sheet['H4'].value + sheet['H5'].value + sheet['H6'].value + sheet['H7'].value + sheet['H8'].value + sheet['H9'].value

    wb.save(year + 'SalesRecord' + '.xlsx')

def pageSetup(wb, sheetName):
    #standard setup for a new sheet
    sheet = wb.get_sheet_by_name(sheetName)
    sheet.freeze_panes = 'A3'
    sheet['A1'].font = Font(name='Times New Roman', bold=True)
    sheet['A2'].font = Font(name='Times New Roman', bold=True)
    sheet['B2'].font = Font(name='Times New Roman', bold=True)
    sheet['C2'].font = Font(name='Times New Roman', bold=True)
    sheet['D2'].font = Font(name='Times New Roman', bold=True)
    sheet['E2'].font = Font(name='Times New Roman', bold=True)
    sheet['F2'].font = Font(name='Times New Roman', bold=True)
    sheet['F1'].number_format = '#,##0 ¥'
    sheet['A1'] = year + '/' + month + ' Sales Record'
    sheet['A2'] = 'Date'
    sheet['B2'] = 'Contract Num'
    sheet['C2'] = 'Purchaser'
    sheet['D2'] = 'Type'
    sheet['E1'] = 'Revenue：'
    sheet['F1'] = 0
    sheet['E2'] = 'Units'
    sheet['F2'] = 'Price'
    default_cells = ['A1', 'A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G1', 'G2', 'G3', 'G4', 'G5', 'G6', 'G7', 'G8', 'G9',
                     'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9']
    sheet['G2'] = 'Total Sale: '
    sheet['G3'] = 'Type A：'
    sheet['G4'] = 'Type B：'
    sheet['G5'] = 'Type BX：'
    sheet['G6'] = 'Type BM：'
    sheet['G7'] = 'Type BS：'
    sheet['G8'] = 'Type XZ：'
    sheet['G9'] = 'Type C：'
    sheet['H3'] = 0
    sheet['H4'] = 0
    sheet['H5'] = 0
    sheet['H6'] = 0
    sheet['H7'] = 0
    sheet['H8'] = 0
    sheet['H9'] = 0
    for c in default_cells:
        sheet[c].border = thin_border
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['A'].width = 15

def makeContractList(path):

    path = path + '\\' + 'InputFile'

    os.chdir(path)
    listOfFile = os.listdir()

    listOfContracts = []

    for i in range(len(listOfFile)):
        listOfContracts.append(Contract.Contract(path + '\\' + listOfFile[i]))

    # x = dict.values()
    #
    # for i in x:
    #     listOfContracts.append(i)
    return listOfContracts



